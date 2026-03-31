import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

thin = Border(
    left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC')
)

def write_sheet(ws, rows, color_hex, col_widths, headers):
    for col, width in col_widths:
        ws.column_dimensions[col].width = width
    hfill = PatternFill('solid', start_color=color_hex)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
    ws.row_dimensions[1].height = 22
    for ri, row in enumerate(rows, 2):
        fill = PatternFill('solid', start_color='F0F9F9') if ri % 2 == 0 else PatternFill('solid', start_color='FFFFFF')
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name='Arial', size=10)
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center' if ci == 1 else 'left', vertical='center')
            cell.border = thin
        ws.row_dimensions[ri].height = 18

students_c1 = [
    ('PRIYA KARTHIK','priya_4095','98374095'),
    ('SOWMIYA VISHWANATHAN','sowmiya_4499','84254499'),
    ('HALITHA PARVEEN S','halitha_3093','92383093'),
    ('RASUL KIFFTHIYA','rasul_4557','67974557'),
    ('SRINITHI','srinithi_9605','89149605'),
    ('YUVASHREE','yuvashree_7136','50267136'),
    ('AFRIN SHIFANA M','afrin_0497','52730497'),
    ('SIRISHA','sirisha_0358','63380358'),
    ('KARISHMA','karishma_9500','95859500'),
    ('REKHA','rekha_2927','48132927'),
    ('SILMIYA RANI','silmiya_3597','74013597'),
    ('SYAMALA','syamala_4098','61534098'),
    ('NAJEEHA NABIL','najeeha_2834','48782834'),
    ('SUDHA N','sudha_7151','84657151'),
    ('DIVYA A','divya_7518','52927518'),
    ('HABIBA','habiba_2424','80042424'),
    ('SWETHA','swetha_5996','00645996'),
    ('ABIRAMI','abirami_7620','40287620'),
    ('SULTHANA','sulthana_9883','83409883'),
    ('SABAANA','sabaana_7003','95897003'),
    ('AYEESHA REEMA','ayeesha_8105','58448105'),
    ('RASIKA S','rasika_7672','58837672'),
    ('GOMATHI','gomathi_2152','26882152'),
    ('FEBINA BEGUM J','febina_9793','83949793'),
    ('PARVEEN BANU A','parveen_1208','11031208'),
    ('KALAIMATHI','kalaimathi_6857','04346857'),
    ('POORNISHA R','poornisha_8268','83348268'),
    ('FATHIMA S','fathima_7217','62417217'),
    ('DHANALAKSHMI S','dhanalakshmi_3732','54343732'),
    ('LATHA A','latha_6078','62886078'),
    ('SHIFANA','shifana_3350','52793350'),
    ('ASMIYA ASFAR','asmiya_2419','00622419'),
    ('JANAGI R','janagi_6801','78906801'),
    ('SAHANA','sahana_1263','00061263'),
    ('MUMTAJ','mumtaj_0533','40380533'),
    ('NITHYA C','nithya_0943','84760943'),
    ('SARALA','sarala_6665','10146665'),
    ('GAYATHRI G','gayathri_3819','20413819'),
    ('MARSUKA BANU B','marsuka_7986','58597986'),
    ('POOVIZHI D','poovizhi_2798','44992798'),
    ('PARVEEN FATHIMA R','parveen_3969','85523969'),
    ('SABINA BEEVI A','sabina_9386','41539386'),
    ('HAMEEDA BEEVI N','hameeda_9910','00099910'),
    ('JAYABHARATHI','jayabharathi_0510','60720510'),
    ('SHALINI','shalini_8330','48378330'),
    ('JANAGI','janagi_1913','72301913'),
    ('NARGIES BANU','nargies_3892','94213892'),
    ('VIDHYASRI S','vidhyasri_8081','84868081'),
    ('JANNATH FIRTHOUSE','jannath_4166','48284166'),
    ('KOWSALYA S','kowsalya_2711','25142711'),
    ('MUTHU KANAGASELVI G','muthu_0555','70420555'),
    ('SHAZIYA M','shaziya_3898','42473898'),
    ('THANZEERA','thanzeera_8172','85858172'),
    ('KALAIVANI M','kalaivani_1527','90581527'),
    ('SHIFANA FATHIMA A','shifana_3003','14043003'),
    ('MAGESHWARI G','mageshwari_1934','45401934'),
    ('AYISHA S','ayisha_1416','85791416'),
    ('GOWRIPRIYA R','gowripriya_7987','67287987'),
    ('MEIVIZHI K','meivizhi_4925','92514925'),
    ('HARSHANA M','harshana_3051','66733051'),
    ('RAHMATH NISHA S','rahmath_9511','03749511'),
    ('GAYATHRI J','gayathri_5391','84935391'),
    ('ASHRATH A','ashrath_0789','44400789'),
    ('PRADEEPA R','pradeepa_4230','84684230'),
    ('SUMAIYA S','sumaiya_8140','40738140'),
    ('SANDHIYA DHAYALAN','sandhiya_2729','00862729'),
    ('GIRIJA SANKAR','girija_0068','94160068'),
    ('KARPAGAM','karpagam_4410','94454410'),
    ('ARCHANA JEYARAM','archana_6420','06306420'),
    ('RAJA KUMARI R','raja_8260','70088260'),
    ('SHREE LEKKA N','shree_0446','20880446'),
    ('AAMINA','aamina_5636','83595636'),
    ('AZMINA BARVEEN','azmina_8005','03368005'),
    ('RUVAITHA ANSARI','ruvaitha_2683','62462683'),
    ('GHOUSIA BEGAM','ghousia_2552','43672552'),
    ('SOBBIYA S','sobbiya_1096','38791096'),
    ('ANUSHEYA','anusheya_5957','05065957'),
    ('POOJA','pooja_4348','76034348'),
    ('JENNIFER DOMINICA CLARKSON','jennifer_0737','38650737'),
    ('JANAKI','janaki_6317','97596317'),
    ('SANDHYA RANI','sandhya_7384','07987384'),
    ('ANJALI DEVI','anjali_3440','78603440'),
    ('PRIYADHARSHINI','priyadharshini_8372','40118372'),
    ('REENA S','reena_8808','74018808'),
    ('TAMILSELVI','tamilselvi_2034','90432034'),
    ('JAYASRI','jayasri_3210','89053210'),
    ('HABEEBA REEM','habeeba_9394','68049394'),
    ('SABINA BEE','sabina_5631','56645631'),
    ('RAHMATH ASMI','rahmath_4461','39454461'),
    ('HEMALAVANYA','hemalavanya_8162','48708162'),
    ('PRAVILIKA','pravilika_4116','10964116'),
]

students_c2 = [
    ('HAMEEDA','hameeda_9910','00099910'),
    ('DIVYA','divya_7738','79377738'),
    ('LOGANAYAGI','loganayagi_0625','49460625'),
]

students_c3 = [
    ('AYISHA','ayisha_2440','89802440'),
    ('VAIRAMANI S','vairamani_6923','86386923'),
    ('RAFIA','rafia_1542','52051542'),
    ('AAMINA','aamina_5636','83595636'),
    ('JAYAPRATHA','jayapratha_3951','62383951'),
    ('SITTA VARALAKSHMI','sitta_2563','63902563'),
    ('ARUNA DEVI','aruna_1886','62531886'),
    ('VIJAYALAKSHMI P','vijayalakshmi_1811','80791811'),
    ('SHIBA','shiba_6329','02246329'),
    ('DIVYAPRASANNA','divyaprasanna_0001','99990001'),
    ('TAMILSELVI','tamilselvi_0002','99990002'),
    ('PADMAPRIYA','padmapriya_0003','99990003'),
]

students_c4 = [
    ('HAFZA','hafza_0221','71930221'),
    ('THASLINISHA','thaslinisha_0557','23510557'),
    ('AMREEN','amreen_4966','45144966'),
    ('FATHIMA','fathima_3158','12883158'),
    ('HAMEETHA PARVEEN','hameetha_2019','18082019'),
    ('KATHIJA','kathija_8121','87008121'),
    ('SIRAJUN NISHA','sirajun_0788','42930788'),
    ('AFRA','afra_1050','87081050'),
    ('SHAHIN','shahin_4872','54884872'),
    ('SUSHMA SULTHANA','sushma_3675','44263675'),
    ('JENNIFER','jennifer_0737','38650737'),
]

staff = [
    ('Mrs. Abinaya','abinaya','abinaya@ansha','Teacher - Montessori'),
    ('Mrs. Shakila Hussain','shakila','shakila@ansha','Teacher - Montessori'),
    ('Mrs. Priya','priya','priya@ansha','Teacher - Phonics'),
    ('Mrs. Syed','syed','syed@ansha','Admin'),
    ('Mrs. Hemala','hemala','hemala@ansha','Admin'),
    ('Mrs. Suvitha','suvitha','suvitha@ansha','Admin'),
]

cols3 = [('A',5),('B',32),('C',24),('D',16)]
h3 = ['S.No','Student Name','Username','Password']

ws1 = wb.active
ws1.title = 'Montessori (c1)'
write_sheet(ws1, [(i+1,n,u,p) for i,(n,u,p) in enumerate(students_c1)], '1a7a7a', cols3, h3)

ws2 = wb.create_sheet('Spoken English (c2)')
write_sheet(ws2, [(i+1,n,u,p) for i,(n,u,p) in enumerate(students_c2)], '7c3aed', cols3, h3)

ws3 = wb.create_sheet('Phonics (c3)')
write_sheet(ws3, [(i+1,n,u,p) for i,(n,u,p) in enumerate(students_c3)], 'd97706', cols3, h3)

ws4 = wb.create_sheet('Child Psychology (c4)')
write_sheet(ws4, [(i+1,n,u,p) for i,(n,u,p) in enumerate(students_c4)], 'db2777', cols3, h3)

ws5 = wb.create_sheet('Teachers & Admins')
cols5 = [('A',5),('B',24),('C',18),('D',18),('E',22)]
h5 = ['S.No','Name','Username','Password','Role']
write_sheet(ws5, [(i+1,n,u,p,r) for i,(n,u,p,r) in enumerate(staff)], '1e3a5f', cols5, h5)

out = r'C:\Users\Syed\Desktop\Montessori\lms\Student_Login_Credentials.xlsx'
wb.save(out)
print('Saved:', out)
